from flask import Flask, jsonify, request, send_from_directory, make_response
from flask_cors import CORS
import os
import logging
import io
import time
import json
import atexit
from typing import List, Dict, Any, Optional
from flaskwebgui import FlaskUI
import analytics

app = Flask(__name__)
CORS(app)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ========== 数据持久化配置 ==========
# 将数据文件保存在当前用户目录下的一个隐藏文件夹中，避免权限问题
DATA_DIR = os.path.join(os.path.expanduser("~"), ".todo_app")
DATA_FILE = os.path.join(DATA_DIR, "todos.json")

def ensure_data_dir():
    """确保数据目录存在"""
    if not os.path.exists(DATA_DIR):
        os.makedirs(DATA_DIR)

def load_todos() -> List[Dict[str, Any]]:
    """从 JSON 文件加载待办事项列表"""
    ensure_data_dir()
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                # 确保每个 todo 都有必要的字段（兼容旧数据）
                for todo in data:
                    if 'priority' not in todo:
                        todo['priority'] = 'mid'
                    if 'due_date' not in todo:
                        todo['due_date'] = None
                    if 'notes' not in todo:
                        todo['notes'] = ''
                    if 'tags' not in todo:
                        todo['tags'] = []
                    if 'steps' not in todo:
                        todo['steps'] = []
                    # 兼容旧 steps 数据
                    for step in todo.get('steps', []):
                        if 'due_date' not in step:
                            step['due_date'] = None
                logger.info(f"从文件加载了 {len(data)} 条待办事项")
                return data
        except Exception as e:
            logger.error(f"加载数据文件失败: {e}，使用默认数据")
    # 无文件或读取失败时返回默认示例数据
    return [
        {"id": 1, "text": "学习Python",    "completed": False, "priority": "high", "due_date": None, "notes": "", "tags": ["学习"], "steps": [
            {"id": 1, "text": "安装Python环境", "completed": True,  "due_date": None},
            {"id": 2, "text": "学习基础语法",   "completed": False, "due_date": None},
            {"id": 3, "text": "完成练习项目",   "completed": False, "due_date": None},
        ]},
        {"id": 2, "text": "开发Flask应用", "completed": True,  "priority": "mid",  "due_date": None, "notes": "", "tags": ["开发"], "steps": []},
        {"id": 3, "text": "学习JavaScript","completed": False, "priority": "low",  "due_date": None, "notes": "", "tags": ["学习"], "steps": []},
    ]

def save_todos(data: List[Dict[str, Any]]):
    """将待办事项列表保存到 JSON 文件"""
    ensure_data_dir()
    try:
        with open(DATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        logger.info(f"数据已保存，当前共 {len(data)} 条")
    except Exception as e:
        logger.error(f"保存数据文件失败: {e}")

# 初始化内存数据（从文件加载）
todos: List[Dict[str, Any]] = load_todos()

# ========== 分析埋点初始化 ==========
analytics.init_db()
analytics.start_session()
atexit.register(analytics.end_session)

# ========== 辅助函数 ==========
def get_next_id() -> int:
    return max((t["id"] for t in todos), default=0) + 1

def find_todo(todo_id: int) -> Optional[Dict[str, Any]]:
    return next((t for t in todos if t["id"] == todo_id), None)

def get_next_step_id(todo: Dict[str, Any]) -> int:
    return max((s["id"] for s in todo.get("steps", [])), default=0) + 1

def validate_due_date(d) -> tuple[bool, str]:
    if d is None:
        return True, ""
    if not isinstance(d, str):
        return False, "due_date 必须是 YYYY-MM-DD 字符串或 null"
    try:
        time.strptime(d, "%Y-%m-%d")
        return True, ""
    except ValueError:
        return False, "due_date 格式错误，应为 YYYY-MM-DD"

VALID_PRIORITIES = {"high", "mid", "low"}

# ========== 宠物状态持久化 ==========
PET_FILE = os.path.join(DATA_DIR, "pet_state.json")

def load_pet() -> Dict[str, Any]:
    """加载宠物状态"""
    ensure_data_dir()
    if os.path.exists(PET_FILE):
        try:
            with open(PET_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    now = time.time()
    return {
        "first_seen": now,
        "character": "egg",
        "unlocked": False,
        "show_pet": True,
        "interactions": 0,
        "last_interact": now,
    }

def save_pet(data: Dict[str, Any]):
    ensure_data_dir()
    try:
        with open(PET_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.error(f"保存宠物状态失败: {e}")

def calculate_rfm_character() -> str:
    """根据待办行为数据（RFM 改编）确定宠物角色，共 9 种（不含蛋）"""
    if not todos:
        return "bear"
    total     = len(todos)
    done_list = [t for t in todos if t.get("completed")]
    done_cnt  = len(done_list)
    pend_cnt  = total - done_cnt
    comp_rate = done_cnt / total if total else 0

    high_cnt  = sum(1 for t in todos if t.get("priority") == "high")
    low_cnt   = sum(1 for t in todos if t.get("priority") == "low")
    has_due   = sum(1 for t in todos if t.get("due_date"))
    has_notes = sum(1 for t in todos if (t.get("notes") or "").strip())
    has_tags  = sum(1 for t in todos if t.get("tags"))
    has_steps = sum(1 for t in todos if t.get("steps"))
    ai_steps  = sum(1 for t in todos if t.get("steps") and len(t.get("steps", [])) >= 3)

    due_rate   = has_due   / total
    notes_rate = has_notes / total
    tags_rate  = has_tags  / total
    steps_rate = has_steps / total
    high_rate  = high_cnt  / total
    low_rate   = low_cnt   / total
    ai_rate    = ai_steps  / total

    # 迷你龙：重度玩家，大量完成 + 高优先级为主
    if total >= 20 and done_cnt >= 15 and high_rate >= 0.45:
        return "dragon"

    # 幽灵喵：极少操作但已完成，神秘稀有
    if total <= 4 and done_cnt >= 2:
        return "ghost"

    # 侦探企鹅：极度有序，大量使用截止日期 + 标签
    if due_rate >= 0.5 and (tags_rate >= 0.4 or steps_rate >= 0.5):
        return "penguin"

    # 智慧狐：善用 AI 与备注，深度规划型
    if (ai_rate + notes_rate) >= 0.4 or (notes_rate >= 0.3 and steps_rate >= 0.3):
        return "fox"

    # 活力兔：高完成量 + 高优先级热衷者
    if done_cnt >= 8 and high_rate >= 0.4:
        return "bunny"

    # 慵懒树懒：偏好低优先级，完成率偏低
    if low_rate >= 0.5 and comp_rate < 0.4:
        return "sloth"

    # 贪睡仓鼠：待办堆积，完成率低
    if pend_cnt >= 8 and comp_rate < 0.35:
        return "hamster"

    # 淡然猫：稳定完成，完成率高
    if comp_rate >= 0.6 and done_cnt >= 5:
        return "cat"

    # 踏实熊：均衡默认
    return "bear"

# ========== 设置持久化 ==========
SETTINGS_FILE = os.path.join(DATA_DIR, "settings.json")

def load_settings() -> Dict[str, Any]:
    """加载应用设置"""
    ensure_data_dir()
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"加载设置失败: {e}")
    return {
        "ai_enabled": False,
        "ai_base_url": "https://api.openai.com/v1",
        "ai_api_key": "",
        "ai_model": "gpt-3.5-turbo",
        "mysql": {
            "host": "localhost",
            "port": 3306,
            "user": "root",
            "password": "159506",
            "database": "todo_analytics"
        }
    }

def save_settings(data: Dict[str, Any]):
    """保存应用设置"""
    ensure_data_dir()
    try:
        with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.error(f"保存设置失败: {e}")

def validate_payload(data: Dict[str, Any]) -> tuple[bool, str]:
    if not data:
        return False, "请求体不能为空"
    if "text" in data:
        text = data["text"]
        if not isinstance(text, str) or not text.strip():
            return False, "text 必须是非空字符串"
        if len(text.strip()) > 500:
            return False, "text 过长（最大 500 字符）"
    if "completed" in data and not isinstance(data["completed"], bool):
        return False, "completed 必须是布尔值"
    if "priority" in data and data["priority"] not in VALID_PRIORITIES:
        return False, f"priority 必须是 {VALID_PRIORITIES} 之一"
    if "notes" in data:
        n = data["notes"]
        if n is not None and not isinstance(n, str):
            return False, "notes 必须是字符串或 null"
        if n and len(n) > 2000:
            return False, "notes 过长（最大 2000 字符）"
    if "tags" in data:
        t = data["tags"]
        if not isinstance(t, list):
            return False, "tags 必须是字符串数组"
        if len(t) > 10:
            return False, "tags 最多 10 个"
        for tag in t:
            if not isinstance(tag, str) or len(tag) > 30:
                return False, "每个 tag 必须是不超过 30 字符的字符串"
    if "due_date" in data:
        d = data["due_date"]
        if d is not None:
            if not isinstance(d, str):
                return False, "due_date 必须是 YYYY-MM-DD 字符串或 null"
            try:
                time.strptime(d, "%Y-%m-%d")
            except ValueError:
                return False, "due_date 格式错误，应为 YYYY-MM-DD"
    return True, ""

# ========== 静态文件 ==========
@app.route("/")
def index():
    return send_from_directory(".", "index.html")

@app.route("/<path:path>")
def serve_static(path):
    if path == "guide.html":
        return jsonify({"code": 404, "message": "Not Found"}), 404
    return send_from_directory(".", path)

# ========== API ==========
@app.route("/api/todos", methods=["GET"])
def get_todos():
    logger.info("GET /api/todos  count=%d", len(todos))
    return jsonify(todos)

@app.route("/api/todos", methods=["POST"])
def add_todo():
    data = request.get_json(silent=True) or {}
    ok, msg = validate_payload(data)
    if not ok:
        return jsonify({"error": msg}), 400
    if "text" not in data:
        return jsonify({"error": "缺少 text 字段"}), 400

    raw_tags = data.get("tags", [])
    clean_tags = [str(tag).strip() for tag in raw_tags if str(tag).strip()]
    todo = {
        "id":        get_next_id(),
        "text":      data["text"].strip(),
        "completed": False,
        "priority":  data.get("priority", "mid"),
        "due_date":  data.get("due_date", None),
        "notes":     (data.get("notes") or "").strip(),
        "tags":      clean_tags,
        "steps":     [],
    }
    todos.append(todo)
    save_todos(todos)          # ← 保存到文件
    analytics.on_todo_created(todo)
    analytics.record_event("todo_created", todo["id"], {
        "priority": todo["priority"],
        "has_due_date": bool(todo["due_date"]),
        "tag_count": len(todo["tags"]),
    })
    logger.info("POST /api/todos  id=%d", todo["id"])
    return jsonify(todo), 201

@app.route("/api/todos/search", methods=["GET"])
def search_todos():
    q = request.args.get("q", "").strip().lower()
    results = todos if not q else [t for t in todos if q in t["text"].lower()]
    if q:
        analytics.record_event("search", extra={"keyword_len": len(q), "results": len(results)})
    logger.info("GET /api/todos/search  q=%r  found=%d", q, len(results))
    return jsonify(results)

@app.route("/api/todos/export", methods=["GET"])
def export_todos():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        return jsonify({"error": "缺少 openpyxl，请执行 pip install openpyxl"}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = "待办事项"

    headers = ["ID", "内容", "优先级", "截止日期", "完成状态", "导出时间"]
    fill = PatternFill("solid", fgColor="C0392B")
    now = time.strftime("%Y-%m-%d %H:%M:%S")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.fill = fill

    pri_labels = {"high": "高", "mid": "中", "low": "低"}

    if todos:
        for row, t in enumerate(todos, 2):
            ws.cell(row=row, column=1, value=t["id"])
            ws.cell(row=row, column=2, value=t["text"])
            ws.cell(row=row, column=3, value=pri_labels.get(t.get("priority","mid"), "中"))
            ws.cell(row=row, column=4, value=t.get("due_date") or "—")
            ws.cell(row=row, column=5, value="已完成" if t["completed"] else "未完成")
            ws.cell(row=row, column=6, value=now)
    else:
        ws.merge_cells("A2:F2")
        ws.cell(row=2, column=1, value="暂无数据").alignment = Alignment(horizontal="center")

    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=8) + 4
        ws.column_dimensions[col[0].column_letter].width = min(w, 60)

    buf = io.BytesIO()
    wb.save(buf)

    resp = make_response(buf.getvalue())
    resp.headers["Content-Type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp.headers["Content-Disposition"] = (
        f'attachment; filename="todos_{time.strftime("%Y%m%d_%H%M%S")}.xlsx"'
    )
    analytics.record_event("export", extra={"count": len(todos)})
    logger.info("GET /api/todos/export  count=%d", len(todos))
    return resp

@app.route("/api/todos/reorder", methods=["PUT"])
def reorder_todos():
    """按客户端传入的 id 列表重新排列 todos"""
    data = request.get_json(silent=True) or {}
    id_order = data.get("order", [])
    if not isinstance(id_order, list):
        return jsonify({"error": "order 必须是 id 数组"}), 400
    id_map = {t["id"]: t for t in todos}
    reordered = [id_map[i] for i in id_order if i in id_map]
    # 补上未在列表中的（安全保底）
    present = set(id_order)
    for t in todos:
        if t["id"] not in present:
            reordered.append(t)
    todos[:] = reordered
    save_todos(todos)
    analytics.record_event("reorder", extra={"count": len(todos)})
    logger.info("PUT /api/todos/reorder  count=%d", len(todos))
    return jsonify({"message": "排序已保存"})

@app.route("/api/stats", methods=["GET"])
def get_stats():
    """返回统计摘要"""
    import time as _t
    today = _t.strftime("%Y-%m-%d")
    total   = len(todos)
    done    = sum(1 for t in todos if t["completed"])
    overdue = sum(1 for t in todos if t.get("due_date") and not t["completed"] and t["due_date"] < today)
    all_tags = {}
    for t in todos:
        for tag in (t.get("tags") or []):
            all_tags[tag] = all_tags.get(tag, 0) + 1
    return jsonify({
        "total": total, "done": done, "pending": total - done,
        "overdue": overdue, "tags": all_tags
    })

@app.route("/api/todos/<int:todo_id>", methods=["PUT"])
def update_todo(todo_id):
    data = request.get_json(silent=True) or {}
    ok, msg = validate_payload(data)
    if not ok:
        return jsonify({"error": msg}), 400

    todo = find_todo(todo_id)
    if not todo:
        return jsonify({"error": "待办事项未找到"}), 404

    prev_completed = todo.get("completed", False)
    if "text"      in data: todo["text"]      = data["text"].strip()
    if "completed" in data: todo["completed"] = data["completed"]
    if "priority"  in data: todo["priority"]  = data["priority"]
    if "due_date"  in data: todo["due_date"]  = data["due_date"]
    if "notes"     in data: todo["notes"]     = (data["notes"] or "").strip()
    if "tags"      in data:
        todo["tags"] = [str(t).strip() for t in data["tags"] if str(t).strip()]

    save_todos(todos)          # ← 保存到文件
    analytics.on_todo_updated(todo, prev_completed)
    if todo.get("completed") and not prev_completed:
        analytics.record_event("todo_completed", todo_id, {
            "priority": todo.get("priority"),
            "has_due_date": bool(todo.get("due_date")),
            "step_count": len(todo.get("steps", [])),
        })
    elif not todo.get("completed") and prev_completed:
        analytics.record_event("todo_uncompleted", todo_id)
    logger.info("PUT /api/todos/%d  %s", todo_id, todo)
    return jsonify(todo)

# ========== 步骤 API ==========
@app.route("/api/todos/<int:todo_id>/steps", methods=["POST"])
def add_step(todo_id):
    todo = find_todo(todo_id)
    if not todo:
        return jsonify({"error": "待办事项未找到"}), 404

    data = request.get_json(silent=True) or {}
    text = (data.get("text") or "").strip()
    if not text:
        return jsonify({"error": "步骤内容不能为空"}), 400
    if len(text) > 200:
        return jsonify({"error": "步骤内容过长（最大 200 字符）"}), 400

    due = data.get("due_date")
    ok, msg = validate_due_date(due)
    if not ok:
        return jsonify({"error": msg}), 400

    if "steps" not in todo:
        todo["steps"] = []
    step = {
        "id":        get_next_step_id(todo),
        "text":      text,
        "completed": False,
        "due_date":  due,
    }
    todo["steps"].append(step)
    save_todos(todos)
    logger.info("POST /api/todos/%d/steps  step_id=%d", todo_id, step["id"])
    return jsonify(step), 201

@app.route("/api/todos/<int:todo_id>/steps/<int:step_id>", methods=["PUT"])
def update_step(todo_id, step_id):
    todo = find_todo(todo_id)
    if not todo:
        return jsonify({"error": "待办事项未找到"}), 404
    step = next((s for s in todo.get("steps", []) if s["id"] == step_id), None)
    if not step:
        return jsonify({"error": "步骤未找到"}), 404

    data = request.get_json(silent=True) or {}
    if "text" in data:
        t = (data["text"] or "").strip()
        if not t:
            return jsonify({"error": "步骤内容不能为空"}), 400
        if len(t) > 200:
            return jsonify({"error": "步骤内容过长"}), 400
        step["text"] = t
    if "completed" in data:
        newly_done = bool(data["completed"]) and not step.get("completed")
        step["completed"] = bool(data["completed"])
        if newly_done:
            analytics.record_event("step_completed", todo_id, {"step_id": step_id})
    if "due_date" in data:
        ok, msg = validate_due_date(data["due_date"])
        if not ok:
            return jsonify({"error": msg}), 400
        step["due_date"] = data["due_date"]

    save_todos(todos)
    logger.info("PUT /api/todos/%d/steps/%d", todo_id, step_id)
    return jsonify(step)

@app.route("/api/todos/<int:todo_id>/steps/<int:step_id>", methods=["DELETE"])
def delete_step(todo_id, step_id):
    todo = find_todo(todo_id)
    if not todo:
        return jsonify({"error": "待办事项未找到"}), 404
    before = len(todo.get("steps", []))
    todo["steps"] = [s for s in todo.get("steps", []) if s["id"] != step_id]
    if len(todo["steps"]) == before:
        return jsonify({"error": "步骤未找到"}), 404
    save_todos(todos)
    logger.info("DELETE /api/todos/%d/steps/%d", todo_id, step_id)
    return jsonify({"message": "步骤已删除"})

@app.route("/api/todos/<int:todo_id>", methods=["DELETE"])
def delete_todo(todo_id):
    target = find_todo(todo_id)
    if not target:
        return jsonify({"error": "待办事项未找到"}), 404

    todos[:] = [t for t in todos if t["id"] != todo_id]
    save_todos(todos)          # ← 保存到文件
    analytics.on_todo_deleted(target)
    analytics.record_event("todo_deleted", todo_id, {
        "was_completed": target.get("completed"),
        "priority": target.get("priority"),
        "had_steps": len(target.get("steps", [])) > 0,
    })
    logger.info("DELETE /api/todos/%d", todo_id)
    return jsonify({"message": "删除成功", "deleted_todo": target})

# ========== 设置 API ==========
@app.route("/api/settings", methods=["GET"])
def get_settings():
    s = load_settings()
    result = dict(s)
    # 脱敏处理 API Key
    key = result.get("ai_api_key", "")
    if key:
        result["ai_api_key"] = key[:4] + "****" + key[-4:] if len(key) > 8 else "****"
    return jsonify(result)

@app.route("/api/settings", methods=["PUT"])
def update_settings():
    data = request.get_json(silent=True) or {}
    s = load_settings()
    if "ai_enabled" in data:
        s["ai_enabled"] = bool(data["ai_enabled"])
    if "ai_base_url" in data:
        s["ai_base_url"] = str(data["ai_base_url"]).strip()
    if "ai_api_key" in data:
        new_key = str(data["ai_api_key"]).strip()
        # 若前端传来脱敏占位符，则保留原 key 不覆盖
        if new_key and "****" not in new_key:
            s["ai_api_key"] = new_key
    if "ai_model" in data:
        s["ai_model"] = str(data["ai_model"]).strip()
    if "mysql" in data and isinstance(data["mysql"], dict):
        m = data["mysql"]
        s.setdefault("mysql", {})
        for key in ("host", "user", "password", "database"):
            if key in m:
                s["mysql"][key] = str(m[key]).strip()
        if "port" in m:
            s["mysql"]["port"] = int(m["port"])
    save_settings(s)
    # MySQL 配置变化时重置连接，让下次自动重连
    analytics._conn = None
    analytics._mysql_available = None
    analytics.init_db()
    logger.info("PUT /api/settings  ai_enabled=%s", s.get("ai_enabled"))
    return jsonify({"message": "设置已保存"})

# ========== 宠物 API ==========
@app.route("/api/pet", methods=["GET"])
def get_pet():
    pet = load_pet()
    now = time.time()
    first_seen = pet.get("first_seen", now)
    days_elapsed = (now - first_seen) / 86400

    # 超过 7 天且还未解锁 → 自动解锁并分配角色
    if not pet.get("unlocked") and days_elapsed >= 7:
        pet["unlocked"] = True
        pet["character"] = calculate_rfm_character()
        save_pet(pet)

    return jsonify({
        "character":   pet.get("character", "egg"),
        "unlocked":    pet.get("unlocked", False),
        "show_pet":    pet.get("show_pet", True),
        "days_elapsed": round(days_elapsed, 1),
        "days_to_unlock": max(0, round(7 - days_elapsed, 1)),
        "interactions": pet.get("interactions", 0),
    })

@app.route("/api/pet/interact", methods=["POST"])
def pet_interact():
    pet = load_pet()
    pet["interactions"] = pet.get("interactions", 0) + 1
    pet["last_interact"] = time.time()
    save_pet(pet)
    return jsonify({"interactions": pet["interactions"]})

@app.route("/api/pet/settings", methods=["PUT"])
def update_pet_settings():
    data = request.get_json(silent=True) or {}
    pet = load_pet()
    if "show_pet" in data:
        pet["show_pet"] = bool(data["show_pet"])
    save_pet(pet)
    return jsonify({"message": "宠物设置已保存"})

# ========== AI 建议步骤 API ==========
@app.route("/api/ai/suggest-steps", methods=["POST"])
def suggest_steps():
    import re
    s = load_settings()
    if not s.get("ai_enabled"):
        return jsonify({"error": "AI 功能未启用，请先在设置中开启并配置 API"}), 400
    if not s.get("ai_api_key"):
        return jsonify({"error": "未配置 API Key，请在设置中填写"}), 400

    data = request.get_json(silent=True) or {}
    goal = (data.get("goal") or "").strip()
    if not goal:
        return jsonify({"error": "目标内容不能为空"}), 400

    try:
        import urllib.request
        import urllib.error

        base_url = s.get("ai_base_url", "https://api.openai.com/v1").rstrip("/")
        model    = s.get("ai_model", "gpt-3.5-turbo")
        api_key  = s.get("ai_api_key", "")

        prompt = (
            f"请为以下目标生成不超过5个具体的执行步骤，每个步骤简洁清晰（20字以内）。\n"
            f"只返回步骤列表，每行一个步骤，不要编号，不要解释，不要多余内容。\n\n"
            f"目标：{goal}"
        )

        payload = json.dumps({
            "model": model,
            "messages": [{"role": "user", "content": prompt}],
            "max_tokens": 300,
            "temperature": 0.7
        }).encode("utf-8")

        req = urllib.request.Request(
            f"{base_url}/chat/completions",
            data=payload,
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {api_key}"
            }
        )

        with urllib.request.urlopen(req, timeout=30) as resp:
            result_json = json.loads(resp.read().decode("utf-8"))

        content = result_json["choices"][0]["message"]["content"].strip()
        steps = [line.strip() for line in content.split("\n") if line.strip()]
        steps = [re.sub(r"^[\d\.\-\*\•、]\s*", "", st).strip() for st in steps]
        steps = [st for st in steps if st][:5]

        # 尝试从请求中获取 todo_id 用于关联分析
        todo_id_for_ai = data.get("todo_id")
        analytics.record_event("ai_suggest_used", todo_id_for_ai, {"steps_returned": len(steps)})
        if todo_id_for_ai:
            analytics.on_ai_steps_used(todo_id_for_ai)
        logger.info("POST /api/ai/suggest-steps  goal=%r  steps=%d", goal, len(steps))
        return jsonify({"steps": steps})

    except Exception as e:
        logger.error(f"AI 建议步骤失败: {e}")
        return jsonify({"error": f"AI 请求失败：{str(e)}"}), 500

# ========== 启动 ==========
if __name__ == '__main__':
    FlaskUI(app=app, server="flask", width=1000, height=700).run()